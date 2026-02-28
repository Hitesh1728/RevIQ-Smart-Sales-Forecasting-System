"""
Power BI Integration Module
Generates .pbix-compatible files and launches Power BI Desktop or Browser
"""

import json
import os
import subprocess
import sys
import platform
import tempfile
import webbrowser
import zipfile
import struct
import logging
from typing import Dict, Any, Optional
import pandas as pd
import numpy as np
from datetime import datetime
import base64

logger = logging.getLogger(__name__)


def _safe_json(obj):
    """JSON serializer that handles numpy types, NaN, Inf, and non-serializable objects."""
    import math
    if isinstance(obj, dict):
        return {str(k): _safe_json(v) for k, v in obj.items() if not isinstance(k, dict)}
    elif isinstance(obj, (list, tuple)):
        return [_safe_json(i) for i in obj]
    elif isinstance(obj, (np.integer,)):
        return int(obj)
    elif isinstance(obj, (np.floating,)):
        v = float(obj)
        if math.isnan(v) or math.isinf(v):
            return 0.0
        return v
    elif isinstance(obj, np.ndarray):
        return [_safe_json(i) for i in obj.tolist()]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return 0.0
        return obj
    elif isinstance(obj, (int, str, bool)) or obj is None:
        return obj
    else:
        return str(obj)


def safe_dumps(obj):
    """Safe json.dumps that handles numpy/NaN/Inf/unhashable types."""
    return json.dumps(_safe_json(obj), ensure_ascii=False)


class PowerBIExporter:
    """
    Exports ML results to Power BI Desktop via:
    1. CSV data files for Power BI to import
    2. A pre-built .pbit template (Power BI Template)
    3. Embedded HTML dashboard that mimics Power BI styling
    4. Attempts to launch Power BI Desktop if installed
    """

    def __init__(self, output_dir: str = "powerbi_output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def export(self, results: Dict[str, Any], filename_prefix: str = "sales_forecast") -> Dict[str, str]:
        """Export data to multiple formats for Power BI"""
        files = {}

        # 1. Export CSVs
        csv_files = self._export_csvs(results, filename_prefix)
        files.update(csv_files)

        # 2. Export Excel file for Power BI (replaces broken .pbit)
        xlsx_path = self._create_pbit_template(results, filename_prefix)
        if xlsx_path:
            files['pbit'] = xlsx_path

        # 3. Create standalone HTML dashboard (Power BI styled)
        html_path = self._create_html_dashboard(results, filename_prefix)
        if html_path:
            files['html_dashboard'] = html_path

        # 4. Create Power BI import instructions
        readme_path = self._create_powerbi_readme(files, filename_prefix)
        files['readme'] = readme_path

        return files

    def _export_csvs(self, results: Dict[str, Any], prefix: str) -> Dict[str, str]:
        files = {}

        # Historical + Predictions CSV
        if 'history' in results and results['history']:
            df_hist = pd.DataFrame(results['history'])
            path = os.path.join(self.output_dir, f"{prefix}_historical.csv")
            df_hist.to_csv(path, index=False)
            files['historical_csv'] = path

        # Forecast CSV
        if 'forecast' in results and results['forecast']:
            df_fore = pd.DataFrame(results['forecast'])
            path = os.path.join(self.output_dir, f"{prefix}_forecast.csv")
            df_fore.to_csv(path, index=False)
            files['forecast_csv'] = path

        # Model Comparison CSV
        if 'model_comparison' in results:
            rows = []
            for model_name, metrics in results['model_comparison'].items():
                row = {'Model': model_name}
                row.update({k: v for k, v in metrics.items() if isinstance(v, (int, float))})
                rows.append(row)
            if rows:
                df_models = pd.DataFrame(rows)
                path = os.path.join(self.output_dir, f"{prefix}_model_comparison.csv")
                df_models.to_csv(path, index=False)
                files['model_comparison_csv'] = path

        # Feature Importance CSV
        if 'feature_importance' in results and results['feature_importance']:
            fi_items = sorted(results['feature_importance'].items(), key=lambda x: x[1], reverse=True)[:20]
            df_fi = pd.DataFrame(fi_items, columns=['Feature', 'Importance'])
            path = os.path.join(self.output_dir, f"{prefix}_feature_importance.csv")
            df_fi.to_csv(path, index=False)
            files['feature_importance_csv'] = path

        # KPI Summary CSV
        summary = results.get('summary', {})
        forecast_summary = results.get('forecast_summary', {})
        metrics = results.get('metrics', {})

        kpi_data = {
            'Metric': ['Total Historical Sales', 'Avg Monthly Sales', 'Max Sales', 'Min Sales',
                       'Sales Growth Rate %', 'Best ML Model', 'Model Accuracy %',
                       'MAE', 'RMSE', 'R2 Score', 'Total Forecast (Next 12M)', 'Avg Monthly Forecast'],
            'Value': [
                summary.get('total_sales', 0),
                summary.get('avg_sales', 0),
                summary.get('max_sales', 0),
                summary.get('min_sales', 0),
                summary.get('growth_rate', 0),
                metrics.get('best_model', 'N/A'),
                metrics.get('accuracy', 0),
                metrics.get('mae', 0),
                metrics.get('rmse', 0),
                metrics.get('r2', 0),
                forecast_summary.get('total_forecast', 0),
                forecast_summary.get('avg_forecast', 0),
            ]
        }
        df_kpi = pd.DataFrame(kpi_data)
        path = os.path.join(self.output_dir, f"{prefix}_kpi_summary.csv")
        df_kpi.to_csv(path, index=False)
        files['kpi_csv'] = path

        return files

    def _create_pbit_template(self, results: Dict[str, Any], prefix: str) -> Optional[str]:
        """
        Creates a properly formatted Excel file (.xlsx) for Power BI import.
        The .pbit format requires proprietary binary structures that cannot be
        reliably generated outside of Power BI Desktop itself.
        This Excel file works perfectly with Power BI via Get Data -> Excel.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side, GradientFill)
            from openpyxl.utils import get_column_letter

            xlsx_path = os.path.join(self.output_dir, f"{prefix}_powerbi_data.xlsx")
            wb = Workbook()

            metrics = results.get('metrics', {})
            summary = results.get('summary', {})
            forecast_summary = results.get('forecast_summary', {})

            # ── Styles ───────────────────────────────────────────────────────
            hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            hdr_fill   = PatternFill('solid', start_color='0078D4')
            best_fill  = PatternFill('solid', start_color='107C10')
            kpi_fill   = PatternFill('solid', start_color='F3F4F6')
            title_font = Font(name='Arial', bold=True, size=13, color='0078D4')
            bold_font  = Font(name='Arial', bold=True, size=10)
            reg_font   = Font(name='Arial', size=10)
            center     = Alignment(horizontal='center', vertical='center')
            left       = Alignment(horizontal='left', vertical='center')
            thin       = Side(style='thin', color='D0D0D0')
            border     = Border(left=thin, right=thin, top=thin, bottom=thin)

            def style_header(cell, is_best=False):
                cell.font = hdr_font
                cell.fill = best_fill if is_best else hdr_fill
                cell.alignment = center
                cell.border = border

            def style_cell(cell, bold=False):
                cell.font = bold_font if bold else reg_font
                cell.alignment = left
                cell.border = border

            def set_col_widths(ws, widths):
                for col, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(col)].width = w

            # ── Sheet 1: KPI Summary ─────────────────────────────────────────
            ws1 = wb.active
            ws1.title = 'KPI Summary'
            ws1.row_dimensions[1].height = 30

            ws1['A1'] = 'AI Sales Forecasting - KPI Summary'
            ws1['A1'].font = title_font
            ws1['A1'].alignment = center
            ws1.merge_cells('A1:C1')

            ws1.append([])
            headers = ['Metric', 'Value', 'Notes']
            ws1.append(headers)
            for i, h in enumerate(headers, 1):
                style_header(ws1.cell(3, i))

            kpis = [
                ('Best ML Model',        metrics.get('best_model', 'N/A').upper(),     'Selected by lowest RMSE'),
                ('Model Accuracy',        f"{metrics.get('accuracy', 0):.1f}%",         '100% - MAPE'),
                ('R2 Score',              f"{metrics.get('r2', 0):.4f}",                'Goodness of fit (1.0 = perfect)'),
                ('MAE',                   f"{metrics.get('mae', 0):,.2f}",              'Mean Absolute Error'),
                ('RMSE',                  f"{metrics.get('rmse', 0):,.2f}",             'Root Mean Squared Error'),
                ('MAPE',                  f"{metrics.get('mape', 0):.2f}%",             'Mean Absolute Percentage Error'),
                ('Total Historical Sales',f"{summary.get('total_sales', 0):,.2f}",      'Sum of all historical records'),
                ('Average Sales',         f"{summary.get('avg_sales', 0):,.2f}",        'Historical average per period'),
                ('Max Sales',             f"{summary.get('max_sales', 0):,.2f}",        'Historical maximum'),
                ('Min Sales',             f"{summary.get('min_sales', 0):,.2f}",        'Historical minimum'),
                ('Growth Rate',           f"{summary.get('growth_rate', 0):+.2f}%",     'First to last period'),
                ('12M Total Forecast',    f"{forecast_summary.get('total_forecast', 0):,.2f}", 'Sum of next 12 forecasted periods'),
                ('12M Avg Forecast',      f"{forecast_summary.get('avg_forecast', 0):,.2f}",   'Average of next 12 forecasted periods'),
            ]
            for row in kpis:
                ws1.append(list(row))
                r = ws1.max_row
                for c in range(1, 4):
                    style_cell(ws1.cell(r, c), bold=(c == 1))
                if row[0] == 'Best ML Model':
                    ws1.cell(r, 2).fill = PatternFill('solid', start_color='DFF6DD')

            set_col_widths(ws1, [28, 20, 38])

            # ── Sheet 2: Historical Data ─────────────────────────────────────
            ws2 = wb.create_sheet('Historical Data')
            history = results.get('history', [])
            if history:
                hist_headers = list(history[0].keys())
                ws2.append(hist_headers)
                for i, h in enumerate(hist_headers, 1):
                    style_header(ws2.cell(1, i))
                for rec in history:
                    row_vals = [rec.get(k, '') for k in hist_headers]
                    ws2.append(row_vals)
                    r = ws2.max_row
                    for c in range(1, len(hist_headers) + 1):
                        style_cell(ws2.cell(r, c))
                set_col_widths(ws2, [16] * len(hist_headers))

            # ── Sheet 3: 12-Month Forecast ───────────────────────────────────
            ws3 = wb.create_sheet('12M Forecast')
            forecast = results.get('forecast', [])
            avg_sales = summary.get('avg_sales', 0)
            if forecast:
                fore_headers = ['#', 'Date', 'Forecasted Value', 'vs Avg Historical', 'vs Avg %', 'Trend']
                ws3.append(fore_headers)
                for i, h in enumerate(fore_headers, 1):
                    style_header(ws3.cell(1, i))
                for idx, rec in enumerate(forecast, 1):
                    val = rec.get('predicted', 0)
                    diff = val - avg_sales
                    diff_pct = (diff / avg_sales * 100) if avg_sales else 0
                    trend = 'UP' if diff >= 0 else 'DOWN'
                    ws3.append([idx, rec.get('date', f'Period+{idx}'),
                                 round(val, 2), round(diff, 2),
                                 f"{diff_pct:+.1f}%", trend])
                    r = ws3.max_row
                    for c in range(1, 7):
                        style_cell(ws3.cell(r, c))
                    if diff >= 0:
                        ws3.cell(r, 6).font = Font(name='Arial', size=10, color='107C10', bold=True)
                    else:
                        ws3.cell(r, 6).font = Font(name='Arial', size=10, color='D13438', bold=True)
                set_col_widths(ws3, [5, 14, 20, 20, 12, 8])

            # ── Sheet 4: Model Comparison ────────────────────────────────────
            ws4 = wb.create_sheet('Model Comparison')
            model_comparison = results.get('model_comparison', {})
            best_model = metrics.get('best_model', '')
            if model_comparison:
                mc_headers = ['Model', 'Accuracy %', 'MAE', 'RMSE', 'R2 Score', 'MAPE %', 'Status']
                ws4.append(mc_headers)
                for i, h in enumerate(mc_headers, 1):
                    style_header(ws4.cell(1, i))
                for name, m in model_comparison.items():
                    if not isinstance(m, dict):
                        continue
                    acc = max(0, 100 - m.get('mape', 100))
                    is_best = name == best_model
                    ws4.append([
                        name.upper(),
                        f"{acc:.1f}%",
                        f"{m.get('mae', 0):.4f}",
                        f"{m.get('rmse', 0):.4f}",
                        f"{m.get('r2', 0):.4f}",
                        f"{m.get('mape', 0):.2f}%",
                        'BEST' if is_best else 'Tested'
                    ])
                    r = ws4.max_row
                    for c in range(1, 8):
                        cell = ws4.cell(r, c)
                        style_cell(cell, bold=is_best)
                        if is_best:
                            cell.fill = PatternFill('solid', start_color='DFF6DD')
                set_col_widths(ws4, [20, 14, 14, 14, 12, 12, 10])

            # ── Sheet 5: Feature Importance ──────────────────────────────────
            ws5 = wb.create_sheet('Feature Importance')
            fi = results.get('feature_importance', {})
            if fi:
                fi_sorted = sorted(fi.items(), key=lambda x: x[1], reverse=True)[:20]
                fi_headers = ['Rank', 'Feature', 'Importance Score']
                ws5.append(fi_headers)
                for i, h in enumerate(fi_headers, 1):
                    style_header(ws5.cell(1, i))
                for rank, (feat, score) in enumerate(fi_sorted, 1):
                    ws5.append([rank, feat, round(float(score), 6)])
                    r = ws5.max_row
                    for c in range(1, 4):
                        style_cell(ws5.cell(r, c), bold=(c == 2))
                set_col_widths(ws5, [8, 30, 20])

            wb.save(xlsx_path)
            logger.info(f"Power BI Excel file created: {xlsx_path}")
            return xlsx_path

        except Exception as e:
            logger.warning(f"Could not create Excel file: {e}")
            return None

    def _create_html_dashboard(self, results: Dict[str, Any], prefix: str) -> str:
        """Create comprehensive 9-section interactive HTML dashboard"""
        from datetime import datetime as dt
        import json as _json

        def safe_dumps(obj):
            try:
                return _json.dumps(obj, default=lambda x: float(x) if hasattr(x,'__float__') else str(x))
            except Exception:
                return '[]'

        metrics          = results.get('metrics', {})
        summary          = results.get('summary', {})
        fc_summary       = results.get('forecast_summary', {})
        profile          = results.get('profile', {})
        history          = results.get('history', [])
        forecast         = results.get('forecast', [])
        model_comparison = results.get('model_comparison', {})
        feature_imp      = results.get('feature_importance', {})
        raw_data         = results.get('raw_data', [])

        # ── Data prep ──────────────────────────────────────────────────────
        hist_dates  = [h.get('date', f'T{i}') for i, h in enumerate(history)]
        hist_actual = [float(h.get('actual', 0)) for h in history]
        hist_pred   = [float(h.get('predicted', 0)) for h in history]
        fore_dates  = [f.get('date', f'F{i}') for i, f in enumerate(forecast)]
        fore_vals   = [float(f.get('predicted', 0)) for f in forecast]

        # Cumulative sales
        cumulative = []
        running = 0
        for v in hist_actual:
            running += v
            cumulative.append(round(running, 2))

        # MoM growth
        mom = []
        for i in range(1, len(hist_actual)):
            prev = hist_actual[i-1]
            curr = hist_actual[i]
            pct = ((curr - prev) / (abs(prev) + 1e-10)) * 100
            mom.append(round(pct, 2))
        mom_dates = hist_dates[1:]

        # Quarterly aggregation
        q_map = {}
        for i, d in enumerate(hist_dates):
            try:
                import datetime
                pd_d = dt.strptime(d[:10], '%Y-%m-%d')
                q_key = f"Q{((pd_d.month-1)//3)+1} {pd_d.year}"
                q_map[q_key] = q_map.get(q_key, 0) + hist_actual[i]
            except Exception:
                pass
        q_labels = list(q_map.keys())
        q_vals   = [round(v, 2) for v in q_map.values()]

        # Seasonal heatmap (month avg)
        m_map = {i: [] for i in range(1, 13)}
        for i, d in enumerate(hist_dates):
            try:
                m = dt.strptime(d[:10], '%Y-%m-%d').month
                m_map[m].append(hist_actual[i])
            except Exception:
                pass
        month_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        seasonal_avg = [round(sum(m_map[m])/max(len(m_map[m]),1), 2) for m in range(1,13)]

        # Forecast slices
        fore_3  = sum(fore_vals[:3])
        fore_6  = sum(fore_vals[:6])
        fore_12 = sum(fore_vals[:12])

        # Error trend
        errors   = [abs(a - p) for a, p in zip(hist_actual, hist_pred)]
        err_pcts = [round(e/(abs(a)+1e-10)*100, 2) for e,a in zip(errors, hist_actual)]

        # Products from feature importance (proxy)
        fi_sorted = sorted(feature_imp.items(), key=lambda x: x[1], reverse=True)[:10]
        fi_names  = [x[0] for x in fi_sorted]
        fi_vals   = [round(x[1], 6) for x in fi_sorted]

        # Safe model comparison
        valid_mc = {k: v for k, v in model_comparison.items() if isinstance(v, dict)}
        mc_safe  = {}
        for k, v in valid_mc.items():
            safe_v = {}
            for kk, vv in v.items():
                if kk in ('model', 'predictions'):
                    continue
                try:
                    safe_v[kk] = float(vv)
                except Exception:
                    safe_v[kk] = str(vv)
            mc_safe[k] = safe_v

        model_names = list(valid_mc.keys())
        model_accs  = [max(0, 100 - v.get('mape', 100)) for v in valid_mc.values()]

        gr = summary.get('growth_rate', 0)
        best_model = metrics.get('best_model', 'N/A').upper()
        accuracy   = metrics.get('accuracy', 0)
        now_str    = dt.now().strftime('%B %d, %Y %H:%M')
        date_str   = dt.now().strftime('%Y-%m-%d')

        # Smart insights
        insights = []
        if gr > 10:
            insights.append(f"📈 Strong growth detected: sales grew by <b>{gr:.1f}%</b> historically.")
        elif gr > 0:
            insights.append(f"📊 Moderate growth of <b>{gr:.1f}%</b> observed in historical data.")
        else:
            insights.append(f"⚠️ Declining trend of <b>{gr:.1f}%</b> — review may be needed.")

        if fore_12 > summary.get('total_sales', 0):
            pct = (fore_12 - summary.get('total_sales',0)) / (abs(summary.get('total_sales',0))+1e-10) * 100
            insights.append(f"🔮 12-month forecast is <b>{pct:.1f}%</b> higher than historical total.")
        else:
            insights.append(f"🔮 12-month forecast total: <b>{fore_12:,.0f}</b>.")

        if accuracy > 90:
            insights.append(f"🏆 Excellent model accuracy: <b>{accuracy:.1f}%</b> using <b>{best_model}</b>.")
        elif accuracy > 75:
            insights.append(f"✅ Good model accuracy: <b>{accuracy:.1f}%</b> using <b>{best_model}</b>.")
        else:
            insights.append(f"⚠️ Model accuracy is <b>{accuracy:.1f}%</b> — more data may improve it.")

        if seasonal_avg:
            peak_m = seasonal_avg.index(max(seasonal_avg))
            low_m  = seasonal_avg.index(min(seasonal_avg))
            insights.append(f"📅 Seasonal peak in <b>{month_names[peak_m]}</b>, lowest in <b>{month_names[low_m]}</b>.")

        if fore_vals:
            next_q_growth = ((fore_vals[2] - hist_actual[-1]) / (abs(hist_actual[-1])+1e-10)) * 100 if hist_actual else 0
            insights.append(f"📆 Next period forecast: <b>{fore_vals[0]:,.0f}</b> (3-month total: <b>{fore_3:,.0f}</b>).")

        if fi_names:
            insights.append(f"🔑 Top sales driver: <b>{fi_names[0]}</b> (importance: {fi_vals[0]:.4f}).")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RevIQ — Sales Intelligence Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#050505;--bg1:#0f0f0f;--bg2:#161616;--bg3:#1f1f1f;
  --border:#252525;--border2:#2f2f2f;
  --text:#ffffff;--muted:#666;--muted2:#999;
  --accent:#c8ff00;--green:#22c55e;--red:#ef4444;
  --blue:#3b82f6;--orange:#f97316;--purple:#a855f7;
}}
body{{font-family:'Times New Roman',Times,serif;background:var(--bg);color:var(--text);min-height:100vh}}

/* ── TOPBAR ── */
.topbar{{background:var(--bg1);border-bottom:1px solid var(--border);padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:50px;position:sticky;top:0;z-index:200}}
.brand{{font-size:16px;font-weight:700;letter-spacing:-0.5px}}
.brand span{{color:var(--accent)}}
.nav{{display:flex;gap:4px}}
.nbtn{{background:transparent;border:1px solid var(--border2);color:var(--muted2);padding:5px 13px;border-radius:5px;cursor:pointer;font-size:11px;font-family:'Times New Roman',serif;transition:all 0.15s;letter-spacing:0.3px}}
.nbtn:hover{{background:var(--bg2);color:var(--text);border-color:#444}}
.nbtn.active{{color:var(--accent);border-color:var(--accent)}}
.nbtn.cta{{background:var(--accent);color:#000;border-color:var(--accent);font-weight:700}}
.nbtn.cta:hover{{background:#b8ef00}}
.topbar-meta{{font-size:11px;color:var(--muted)}}
.topbar-meta strong{{color:var(--green)}}

/* ── HERO ── */
.hero{{background:var(--bg1);border-bottom:1px solid var(--border);padding:22px 24px;position:relative;overflow:hidden}}
.hero::after{{content:'';position:absolute;top:-80px;right:-80px;width:350px;height:350px;background:radial-gradient(circle,rgba(200,255,0,0.04) 0%,transparent 70%);border-radius:50%;pointer-events:none}}
.hero-row{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px}}
.hero-title{{font-size:24px;font-weight:700;letter-spacing:-0.5px}}
.hero-title em{{color:var(--accent);font-style:normal}}
.hero-sub{{font-size:11px;color:var(--muted);margin-top:3px;letter-spacing:0.5px}}
.report-meta{{text-align:right;font-size:11px;color:var(--muted);line-height:2}}
.report-meta strong{{color:var(--text)}}

/* ── KPI GRID ── */
.kpi-grid{{display:grid;grid-template-columns:repeat(8,1fr);gap:1px;background:var(--border);border:1px solid var(--border);border-radius:10px;overflow:hidden}}
.kpi{{background:var(--bg1);padding:14px 16px;position:relative;transition:background 0.15s;cursor:default}}
.kpi:hover{{background:var(--bg2)}}
.kpi-tag{{font-size:9px;text-transform:uppercase;letter-spacing:2px;color:var(--muted);margin-bottom:5px}}
.kpi-val{{font-size:22px;font-weight:700;line-height:1;font-variant-numeric:tabular-nums}}
.kpi-sub{{font-size:10px;color:var(--muted);margin-top:3px}}
.ka .kpi-val{{color:var(--accent)}}
.kg .kpi-val{{color:var(--green)}}
.kr .kpi-val{{color:var(--red)}}
.kb .kpi-val{{color:var(--blue)}}
.ko .kpi-val{{color:var(--orange)}}
.kp .kpi-val{{color:var(--purple)}}
.kbadge{{position:absolute;top:10px;right:10px;font-size:9px;padding:2px 7px;border-radius:20px;font-weight:700}}
.kbadge-a{{background:rgba(200,255,0,0.12);color:var(--accent)}}
.kbadge-g{{background:rgba(34,197,94,0.12);color:var(--green)}}
.kbadge-r{{background:rgba(239,68,68,0.12);color:var(--red)}}

/* ── LAYOUT ── */
.wrap{{padding:20px 24px;max-width:1700px;margin:0 auto}}
.sec-title{{font-size:10px;text-transform:uppercase;letter-spacing:3px;color:var(--muted);border-bottom:1px solid var(--border);padding-bottom:7px;margin:24px 0 14px;display:flex;align-items:center;gap:8px}}
.sec-title .num{{color:var(--accent)}}
.grid2{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px}}
.grid3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:12px}}
.grid4{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-bottom:12px}}
.grid13{{display:grid;grid-template-columns:2fr 1fr;gap:12px;margin-bottom:12px}}
.grid31{{display:grid;grid-template-columns:3fr 1fr;gap:12px;margin-bottom:12px}}

/* ── CARDS ── */
.card{{background:var(--bg1);border:1px solid var(--border);border-radius:8px;padding:16px;position:relative}}
.card-title{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:var(--muted);margin-bottom:12px;display:flex;align-items:center;gap:6px}}
.card-title .ico{{font-size:13px}}
.card.mb{{margin-bottom:12px}}

/* ── TABS ── */
.tab-bar{{display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:18px;overflow-x:auto}}
.tab{{padding:10px 18px;font-size:11px;cursor:pointer;border:none;background:transparent;color:var(--muted);font-family:'Times New Roman',serif;letter-spacing:0.5px;border-bottom:2px solid transparent;transition:all 0.15s;white-space:nowrap}}
.tab.active{{color:var(--accent);border-bottom-color:var(--accent)}}
.tab:hover:not(.active){{color:var(--text)}}
.panel{{display:none}}.panel.active{{display:block}}

/* ── TABLES ── */
.tbl{{width:100%;border-collapse:collapse;font-size:11px}}
.tbl th{{font-size:9px;text-transform:uppercase;letter-spacing:1.5px;color:var(--muted);padding:8px 10px;border-bottom:1px solid var(--border);text-align:left;background:var(--bg);white-space:nowrap}}
.tbl td{{padding:8px 10px;border-bottom:1px solid var(--border2);color:var(--text);vertical-align:middle}}
.tbl tr:last-child td{{border-bottom:none}}
.tbl tr:hover td{{background:var(--bg2)}}
.tbl .best-row td{{color:var(--accent);font-weight:700}}
.badge{{display:inline-block;padding:2px 8px;border-radius:4px;font-size:9px;font-weight:700;letter-spacing:0.5px}}
.badge-a{{background:rgba(200,255,0,0.12);color:var(--accent)}}
.badge-g{{background:rgba(34,197,94,0.12);color:var(--green)}}
.badge-r{{background:rgba(239,68,68,0.12);color:var(--red)}}
.badge-b{{background:rgba(59,130,246,0.12);color:var(--blue)}}

/* ── PROGRESS ── */
.prog{{height:4px;background:var(--bg3);border-radius:2px;overflow:hidden;margin-top:5px}}
.prog-fill{{height:100%;background:var(--accent);border-radius:2px}}
.prog.blue .prog-fill{{background:var(--blue)}}
.prog.green .prog-fill{{background:var(--green)}}
.prog.red .prog-fill{{background:var(--red)}}

/* ── INSIGHT BOXES ── */
.insights-grid{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:12px}}
.insight{{background:var(--bg1);border:1px solid var(--border);border-left:3px solid var(--accent);border-radius:6px;padding:12px 14px;font-size:12px;line-height:1.7;color:var(--muted2)}}
.insight.green{{border-left-color:var(--green)}}
.insight.red{{border-left-color:var(--red)}}
.insight.blue{{border-left-color:var(--blue)}}
.insight.orange{{border-left-color:var(--orange)}}
.insight b{{color:var(--text)}}

/* ── DRILLTHROUGH ── */
.drill-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px}}
.drill-card{{background:var(--bg1);border:1px solid var(--border);border-radius:8px;padding:14px;cursor:pointer;transition:all 0.15s}}
.drill-card:hover{{border-color:var(--accent);background:var(--bg2)}}
.drill-card.selected{{border-color:var(--accent);box-shadow:0 0 0 1px var(--accent)}}
.drill-name{{font-size:11px;font-weight:700;color:var(--text);margin-bottom:6px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
.drill-val{{font-size:18px;font-weight:700;color:var(--accent)}}
.drill-sub{{font-size:10px;color:var(--muted);margin-top:2px}}

/* ── HEATMAP ── */
.heatmap{{display:grid;grid-template-columns:repeat(12,1fr);gap:4px}}
.heat-cell{{border-radius:4px;padding:8px 4px;text-align:center;font-size:10px;font-weight:700;transition:all 0.2s}}
.heat-label{{font-size:9px;color:var(--muted);text-align:center;margin-top:3px}}

/* ── FORECAST TOGGLE ── */
.fc-toggle{{display:flex;gap:6px;margin-bottom:14px}}
.fc-btn{{background:var(--bg2);border:1px solid var(--border2);color:var(--muted2);padding:6px 16px;border-radius:5px;cursor:pointer;font-size:11px;font-family:'Times New Roman',serif;transition:all 0.15s}}
.fc-btn.active{{background:rgba(200,255,0,0.1);border-color:var(--accent);color:var(--accent)}}

/* ── REPORT BOX ── */
.report-box{{background:var(--bg1);border:1px solid var(--border);border-radius:8px;padding:20px 24px;line-height:1.9;font-size:12px;color:var(--muted2);margin-bottom:12px}}
.report-box h3{{font-size:13px;font-weight:700;color:var(--text);margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid var(--border)}}
.report-box p{{margin-bottom:8px}}
.report-box strong{{color:var(--accent)}}
.report-box ul{{padding-left:18px;margin-bottom:8px}}
.report-box li{{margin-bottom:4px}}

/* ── MINI KPI in cards ── */
.mini-kpi{{display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid var(--border2)}}
.mini-kpi:last-child{{border-bottom:none}}
.mini-kpi-label{{font-size:11px;color:var(--muted2)}}
.mini-kpi-val{{font-size:13px;font-weight:700;color:var(--text)}}

/* ── FOOTER ── */
.footer{{background:var(--bg1);border-top:1px solid var(--border);padding:12px 24px;display:flex;justify-content:space-between;align-items:center;font-size:10px;color:var(--muted);margin-top:28px}}

/* ── PRINT ── */
@media print{{
  .topbar,.nav{{display:none!important}}
  body{{background:#fff;color:#111}}
  .card,.kpi,.report-box{{background:#fff!important;border-color:#ddd!important}}
  :root{{--bg:#fff;--bg1:#fff;--bg2:#f9f9f9;--text:#111;--muted:#555;--border:#ddd}}
}}
::-webkit-scrollbar{{width:4px;height:4px}}
::-webkit-scrollbar-track{{background:var(--bg)}}
::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:2px}}
</style>
</head>
<body>

<!-- TOPBAR -->
<div class="topbar">
  <div class="brand">⚡ Rev<span>IQ</span> <span style="color:var(--muted);font-weight:400;font-size:13px">/ Sales Intelligence</span></div>
  <div class="nav">
    <button class="nbtn active" onclick="showTab('overview')">Overview</button>
    <button class="nbtn" onclick="showTab('sales')">Sales Trend</button>
    <button class="nbtn" onclick="showTab('products')">Products</button>
    <button class="nbtn" onclick="showTab('forecast')">Forecast</button>
    <button class="nbtn" onclick="showTab('time')">Time Intel</button>
    <button class="nbtn" onclick="showTab('models')">AI Models</button>
    <button class="nbtn" onclick="showTab('report')">Report</button>
    <button class="nbtn" onclick="window.print()">🖨</button>
    <button class="nbtn" onclick="window.history.length>1?window.history.back():window.close()" style="color:var(--muted2)">← Back</button>
    <button class="nbtn cta" onclick="saveHTML()">↓ Save</button>
  </div>
  <div class="topbar-meta">
    Best: <strong>{best_model}</strong> &nbsp;·&nbsp; Accuracy: <strong>{accuracy:.1f}%</strong>
    &nbsp;·&nbsp; {now_str}
  </div>
</div>

<!-- HERO -->
<div class="hero">
  <div class="hero-row">
    <div>
      <div class="hero-title">Sales <em>Intelligence</em> Dashboard</div>
      <div class="hero-sub">AutoML · 6 Models · 12-Month Forecast · Generated by RevIQ</div>
    </div>
    <div class="report-meta">
      <div>Dataset: <strong>{profile.get('rows',0):,} records</strong></div>
      <div>Target: <strong>{profile.get('target_column','N/A')}</strong></div>
      <div>Date col: <strong>{profile.get('date_column','N/A')}</strong></div>
      <div>Generated: <strong>{date_str}</strong></div>
    </div>
  </div>

  <!-- KPI GRID - 8 cards -->
  <div class="kpi-grid">
    <div class="kpi ka">
      <div class="kpi-tag">Total Sales</div>
      <div class="kpi-val">{summary.get('total_sales',0):,.0f}</div>
      <div class="kpi-sub">{len(history)} periods</div>
      <span class="kbadge kbadge-a">TOTAL</span>
    </div>
    <div class="kpi {'kg' if gr>=0 else 'kr'}">
      <div class="kpi-tag">Sales Growth</div>
      <div class="kpi-val">{gr:+.1f}%</div>
      <div class="kpi-sub">{'▲ Growing' if gr>=0 else '▼ Declining'}</div>
      <span class="kbadge {'kbadge-g' if gr>=0 else 'kbadge-r'}">{'GROWTH' if gr>=0 else 'DECLINE'}</span>
    </div>
    <div class="kpi ko">
      <div class="kpi-tag">3M Forecast</div>
      <div class="kpi-val">{fore_3:,.0f}</div>
      <div class="kpi-sub">Next 3 months</div>
    </div>
    <div class="kpi ko">
      <div class="kpi-tag">6M Forecast</div>
      <div class="kpi-val">{fore_6:,.0f}</div>
      <div class="kpi-sub">Next 6 months</div>
    </div>
    <div class="kpi ka">
      <div class="kpi-tag">12M Forecast</div>
      <div class="kpi-val">{fore_12:,.0f}</div>
      <div class="kpi-sub">Full year ahead</div>
      <span class="kbadge kbadge-a">FORECAST</span>
    </div>
    <div class="kpi kg">
      <div class="kpi-tag">Forecast Accuracy</div>
      <div class="kpi-val">{accuracy:.1f}%</div>
      <div class="kpi-sub">{best_model}</div>
      <span class="kbadge kbadge-g">BEST MODEL</span>
    </div>
    <div class="kpi kb">
      <div class="kpi-tag">Avg per Period</div>
      <div class="kpi-val">{summary.get('avg_sales',0):,.0f}</div>
      <div class="kpi-sub">Historical avg</div>
    </div>
    <div class="kpi">
      <div class="kpi-tag">R² Score</div>
      <div class="kpi-val">{metrics.get('r2',0):.3f}</div>
      <div class="kpi-sub">MAE: {metrics.get('mae',0):,.1f}</div>
    </div>
  </div>
</div>

<div class="wrap">
<div class="tab-bar">
  <button class="tab active" id="tab-overview"  onclick="showTab('overview')">📊 Overview</button>
  <button class="tab"        id="tab-sales"     onclick="showTab('sales')">📈 Sales Trend</button>
  <button class="tab"        id="tab-products"  onclick="showTab('products')">🏆 Products</button>
  <button class="tab"        id="tab-forecast"  onclick="showTab('forecast')">🔮 Forecast</button>
  <button class="tab"        id="tab-time"      onclick="showTab('time')">🕒 Time Intel</button>
  <button class="tab"        id="tab-models"    onclick="showTab('models')">🤖 AI Models</button>
  <button class="tab"        id="tab-report"    onclick="showTab('report')">📋 Report</button>
</div>

<!-- ════════════════════════════════════════════════════════════ OVERVIEW -->
<div class="panel active" id="panel-overview">

  <div class="sec-title"><span class="num">01</span> Smart Business Insights</div>
  <div class="insights-grid" id="insightsGrid"></div>

  <div class="sec-title"><span class="num">02</span> Sales Overview</div>
  <div class="card mb">
    <div class="card-title"><span class="ico">📈</span>Actual vs Predicted + 12-Month Forecast</div>
    <canvas id="mainChart" style="max-height:320px"></canvas>
  </div>

  <div class="grid2">
    <div class="card">
      <div class="card-title"><span class="ico">📊</span>Forecast vs Actual (Clustered)</div>
      <canvas id="clusteredChart" style="max-height:240px"></canvas>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">📉</span>Error Trend Over Time</div>
      <canvas id="errorChart" style="max-height:240px"></canvas>
    </div>
  </div>

</div>

<!-- ════════════════════════════════════════════════════════ SALES TREND -->
<div class="panel" id="panel-sales">

  <div class="sec-title"><span class="num">03</span> Sales Trend Analysis</div>
  <div class="card mb">
    <div class="card-title"><span class="ico">📈</span>Actual Sales Over Time</div>
    <canvas id="actualChart" style="max-height:280px"></canvas>
  </div>
  <div class="grid2">
    <div class="card">
      <div class="card-title"><span class="ico">📐</span>Cumulative Sales</div>
      <canvas id="cumulChart" style="max-height:240px"></canvas>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">📉</span>Residuals (Prediction Error)</div>
      <canvas id="residChart" style="max-height:240px"></canvas>
    </div>
  </div>

</div>

<!-- ═══════════════════════════════════════════════════════════ PRODUCTS -->
<div class="panel" id="panel-products">

  <div class="sec-title"><span class="num">04</span> Product / Feature Performance</div>

  <div class="grid2">
    <div class="card">
      <div class="card-title"><span class="ico">🥇</span>Top Features by Sales Impact</div>
      <canvas id="topFeatChart" style="max-height:280px"></canvas>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">📊</span>Pareto — Contribution % (80/20)</div>
      <canvas id="paretoChart" style="max-height:280px"></canvas>
    </div>
  </div>

  <div class="sec-title"><span class="num">05</span> Drillthrough — Click a Feature to Analyse</div>
  <div class="drill-grid" id="drillGrid"></div>
  <div class="grid2" id="drillDetail" style="display:none">
    <div class="card">
      <div class="card-title"><span class="ico">🔍</span>Feature: <span id="drillName">—</span></div>
      <div id="drillStats"></div>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">📊</span>Importance Breakdown</div>
      <canvas id="drillChart" style="max-height:220px"></canvas>
    </div>
  </div>

</div>

<!-- ═══════════════════════════════════════════════════════════ FORECAST -->
<div class="panel" id="panel-forecast">

  <div class="sec-title"><span class="num">06</span> Forecast vs Actual Analysis</div>

  <div class="fc-toggle">
    <button class="fc-btn active" onclick="setFcPeriod(3,this)">3 Months</button>
    <button class="fc-btn" onclick="setFcPeriod(6,this)">6 Months</button>
    <button class="fc-btn" onclick="setFcPeriod(12,this)">12 Months</button>
  </div>

  <div class="card mb">
    <div class="card-title"><span class="ico">🔮</span>Sales Forecast — <span id="fcPeriodLabel">3 Months</span></div>
    <canvas id="fcBarChart" style="max-height:280px"></canvas>
  </div>

  <div class="grid3">
    <div class="kpi ko" style="border-radius:8px;border:1px solid var(--border)">
      <div class="kpi-tag">3M Forecast Total</div>
      <div class="kpi-val" style="font-size:18px">{fore_3:,.0f}</div>
      <div class="kpi-sub">Periods 1–3</div>
    </div>
    <div class="kpi kb" style="border-radius:8px;border:1px solid var(--border)">
      <div class="kpi-tag">6M Forecast Total</div>
      <div class="kpi-val" style="font-size:18px">{fore_6:,.0f}</div>
      <div class="kpi-sub">Periods 1–6</div>
    </div>
    <div class="kpi ka" style="border-radius:8px;border:1px solid var(--border)">
      <div class="kpi-tag">12M Forecast Total</div>
      <div class="kpi-val" style="font-size:18px">{fore_12:,.0f}</div>
      <div class="kpi-sub">Full year ahead</div>
    </div>
  </div>

  <div class="card mt" style="margin-top:12px">
    <div class="card-title"><span class="ico">📋</span>Forecast Detail Table</div>
    <table class="tbl">
      <thead><tr><th>#</th><th>Period</th><th>Forecast</th><th>vs Avg</th><th>Δ%</th><th>Trend</th><th>Bar</th></tr></thead>
      <tbody id="fcTbody"></tbody>
    </table>
  </div>

</div>

<!-- ═══════════════════════════════════════════════════════ TIME INTEL -->
<div class="panel" id="panel-time">

  <div class="sec-title"><span class="num">07</span> Time Intelligence Analysis</div>

  <div class="grid2">
    <div class="card">
      <div class="card-title"><span class="ico">📅</span>Month-over-Month Growth %</div>
      <canvas id="momChart" style="max-height:240px"></canvas>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">📆</span>Quarterly Sales Trend</div>
      <canvas id="quarterChart" style="max-height:240px"></canvas>
    </div>
  </div>

  <div class="sec-title"><span class="num">08</span> Seasonal Heatmap</div>
  <div class="card mb">
    <div class="card-title"><span class="ico">🌡</span>Average Sales by Month (Seasonality)</div>
    <div class="heatmap" id="heatmapGrid"></div>
    <div class="heatmap" id="heatmapLabels" style="margin-top:4px"></div>
  </div>

</div>

<!-- ════════════════════════════════════════════════════════ AI MODELS -->
<div class="panel" id="panel-models">

  <div class="sec-title"><span class="num">09</span> AI Model Performance</div>

  <div class="grid13">
    <div class="card">
      <div class="card-title"><span class="ico">🤖</span>Model Accuracy Comparison</div>
      <canvas id="modelBarChart" style="max-height:260px"></canvas>
    </div>
    <div class="card">
      <div class="card-title"><span class="ico">🏆</span>Best Model</div>
      <div style="text-align:center;padding:20px 0">
        <div style="font-size:36px;margin-bottom:8px">🏆</div>
        <div style="font-size:20px;font-weight:700;color:var(--accent)">{best_model}</div>
        <div style="font-size:13px;color:var(--green);margin-top:6px">{accuracy:.1f}% Accuracy</div>
        <div style="font-size:11px;color:var(--muted);margin-top:4px">R² = {metrics.get('r2',0):.4f}</div>
        <div style="font-size:11px;color:var(--muted)">RMSE = {metrics.get('rmse',0):,.2f}</div>
        <div style="font-size:11px;color:var(--muted)">MAE = {metrics.get('mae',0):,.2f}</div>
      </div>
    </div>
  </div>

  <div class="card mb">
    <div class="card-title"><span class="ico">📊</span>Model Comparison Table</div>
    <table class="tbl">
      <thead><tr><th>Model</th><th>Accuracy</th><th>MAE</th><th>RMSE</th><th>R²</th><th>MAPE</th><th>Accuracy Bar</th><th>Status</th></tr></thead>
      <tbody id="modelTbody"></tbody>
    </table>
  </div>

  <div class="card mb">
    <div class="card-title"><span class="ico">🔑</span>Feature Importance (Top Predictors)</div>
    <canvas id="fiChart" style="max-height:280px"></canvas>
  </div>

</div>

<!-- ════════════════════════════════════════════════════════════ REPORT -->
<div class="panel" id="panel-report">

  <div class="sec-title"><span class="num">10</span> Executive Summary Report</div>

  <div class="report-box">
    <h3>📋 Executive Summary</h3>
    <p>Report generated by <strong>RevIQ AI Sales Forecasting System</strong> on <strong>{now_str}</strong>.</p>
    <p>Dataset: <strong>{profile.get('rows',0):,} records</strong> · Target: <strong>{profile.get('target_column','N/A')}</strong> · Date: <strong>{profile.get('date_column','N/A')}</strong></p>
    <p>AutoML trained <strong>6 models</strong> and selected <strong>{best_model}</strong> with <strong>{accuracy:.1f}%</strong> accuracy (R²={metrics.get('r2',0):.4f}).</p>
    <p>Historical total: <strong>{summary.get('total_sales',0):,.0f}</strong> · Average: <strong>{summary.get('avg_sales',0):,.0f}</strong>/period · Growth: <strong>{gr:+.1f}%</strong></p>
  </div>

  <div class="report-box">
    <h3>🔮 Forecast Outlook</h3>
    <ul>
      <li>3-Month Forecast: <strong>{fore_3:,.0f}</strong></li>
      <li>6-Month Forecast: <strong>{fore_6:,.0f}</strong></li>
      <li>12-Month Forecast: <strong>{fore_12:,.0f}</strong> (avg {fc_summary.get('avg_forecast',0):,.0f}/period)</li>
      <li>Peak forecast period: <strong>{fc_summary.get('max_forecast',0):,.0f}</strong></li>
      <li>Lowest forecast period: <strong>{fc_summary.get('min_forecast',0):,.0f}</strong></li>
    </ul>
  </div>

  <div class="report-box">
    <h3>🤖 Model Performance</h3>
    <p>Best model: <strong>{best_model}</strong> · Accuracy: <strong>{accuracy:.1f}%</strong> · MAE: <strong>{metrics.get('mae',0):,.2f}</strong> · RMSE: <strong>{metrics.get('rmse',0):,.2f}</strong> · MAPE: <strong>{metrics.get('mape',0):.2f}%</strong></p>
    <p>6 competing models: XGBoost, LightGBM, Random Forest, Gradient Boosting, Extra Trees, Ridge.</p>
  </div>

  <div class="report-box">
    <h3>💡 Recommendations</h3>
    <ul id="recList"></ul>
  </div>

  <div style="display:flex;gap:10px;margin-top:16px">
    <button class="nbtn cta" onclick="window.print()" style="padding:10px 22px;font-size:12px">🖨 Print Report</button>
    <button class="nbtn" onclick="saveHTML()" style="padding:10px 22px;font-size:12px">↓ Save HTML</button>
    <button class="nbtn" onclick="exportCSV()" style="padding:10px 22px;font-size:12px">↓ Export Forecast CSV</button>
  </div>

</div>
</div><!-- /wrap -->

<div class="footer">
  <span>⚡ RevIQ Sales Intelligence · AutoML Engine · {dt.now().year}</span>
  <span>Best: <strong style="color:var(--accent)">{best_model}</strong> · Accuracy: <strong style="color:var(--green)">{accuracy:.1f}%</strong> · {profile.get('rows',0):,} records</span>
</div>

<script>
// ── RAW DATA ──────────────────────────────────────────────────────────────────
const histDates   = {safe_dumps(hist_dates)};
const histActual  = {safe_dumps(hist_actual)};
const histPred    = {safe_dumps(hist_pred)};
const foreDates   = {safe_dumps(fore_dates)};
const foreVals    = {safe_dumps(fore_vals)};
const cumulative  = {safe_dumps(cumulative)};
const momDates    = {safe_dumps(mom_dates)};
const momVals     = {safe_dumps(mom)};
const qLabels     = {safe_dumps(q_labels)};
const qVals       = {safe_dumps(q_vals)};
const seasonalAvg = {safe_dumps(seasonal_avg)};
const errPcts     = {safe_dumps(err_pcts)};
const fiNames     = {safe_dumps(fi_names)};
const fiVals      = {safe_dumps(fi_vals)};
const modelComp   = {safe_dumps(mc_safe)};
const avgSales    = {summary.get('avg_sales', 0)};
const bestModel   = "{metrics.get('best_model','xgboost')}";
const insights    = {safe_dumps(insights)};
const MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

const AC = '#c8ff00', GR = '#22c55e', RD = '#ef4444', BL = '#3b82f6', OR = '#f97316';
const GRID = '#1f1f1f', TXT = '#666', FONT = "'Times New Roman'";

const BASE = {{
  responsive:true, maintainAspectRatio:true,
  plugins:{{legend:{{labels:{{color:TXT,font:{{family:FONT,size:10}}}}}}}},
  scales:{{
    x:{{ticks:{{color:TXT,font:{{family:FONT,size:9}}}},grid:{{color:GRID}}}},
    y:{{ticks:{{color:TXT,font:{{family:FONT,size:9}},callback:v=>v.toLocaleString()}},grid:{{color:GRID}}}}
  }}
}};
function deepMerge(a,b){{return Object.assign({{}},a,b,{{scales:Object.assign({{}},a.scales,b.scales)}})}}

// ── CHARTS ────────────────────────────────────────────────────────────────────

// 1. MAIN CHART
new Chart(document.getElementById('mainChart'),{{
  type:'line',
  data:{{
    labels:[...histDates,...foreDates],
    datasets:[
      {{label:'Actual',data:[...histActual,...Array(foreDates.length).fill(null)],
        borderColor:'#fff',backgroundColor:'rgba(255,255,255,0.04)',borderWidth:1.5,pointRadius:0,fill:true,tension:0.3}},
      {{label:'Predicted',data:[...histPred,...Array(foreDates.length).fill(null)],
        borderColor:'#444',backgroundColor:'transparent',borderWidth:1.5,borderDash:[4,3],pointRadius:0,tension:0.3}},
      {{label:'Forecast',data:[...Array(histDates.length).fill(null),...foreVals],
        borderColor:AC,backgroundColor:'rgba(200,255,0,0.06)',borderWidth:2,pointRadius:3,pointBackgroundColor:AC,fill:true,tension:0.3}}
    ]
  }},
  options:deepMerge(BASE,{{plugins:{{legend:{{position:'top',labels:{{color:'#888',font:{{family:FONT}}}}}}}},
    scales:{{x:{{ticks:{{maxTicksLimit:14}}}},y:{{}}}} }})
}});

// 2. CLUSTERED ACTUAL VS PREDICTED
const sampleStep = Math.max(1, Math.floor(histActual.length/30));
const clLabels = histDates.filter((_,i)=>i%sampleStep===0);
const clActual = histActual.filter((_,i)=>i%sampleStep===0);
const clPred   = histPred.filter((_,i)=>i%sampleStep===0);
new Chart(document.getElementById('clusteredChart'),{{
  type:'bar',
  data:{{labels:clLabels,datasets:[
    {{label:'Actual',data:clActual,backgroundColor:'rgba(255,255,255,0.15)',borderRadius:3}},
    {{label:'Predicted',data:clPred,backgroundColor:'rgba(200,255,0,0.4)',borderRadius:3}}
  ]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{position:'top'}}}},scales:{{x:{{ticks:{{maxRotation:45,font:{{size:8}}}}}},y:{{}}}}  }})
}});

// 3. ERROR TREND
new Chart(document.getElementById('errorChart'),{{
  type:'line',
  data:{{labels:histDates,datasets:[{{
    label:'Error %',data:errPcts,borderColor:RD,backgroundColor:'rgba(239,68,68,0.08)',
    borderWidth:1.5,pointRadius:0,fill:true,tension:0.4
  }}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},
    scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y:{{ticks:{{callback:v=>v+'%'}}}}}}  }})
}});

// 4. ACTUAL ONLY
new Chart(document.getElementById('actualChart'),{{
  type:'line',
  data:{{labels:histDates,datasets:[{{
    label:'Actual Sales',data:histActual,borderColor:'#fff',
    backgroundColor:'rgba(255,255,255,0.05)',borderWidth:2,pointRadius:0,fill:true,tension:0.3
  }}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:14}}}},y:{{}}}}  }})
}});

// 5. CUMULATIVE
new Chart(document.getElementById('cumulChart'),{{
  type:'line',
  data:{{labels:histDates,datasets:[{{
    label:'Cumulative Sales',data:cumulative,borderColor:AC,
    backgroundColor:'rgba(200,255,0,0.06)',borderWidth:2,pointRadius:0,fill:true,tension:0.3
  }}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:12}}}},y:{{}}}}  }})
}});

// 6. RESIDUALS SCATTER
const residData=histActual.map((a,i)=>({{x:histPred[i],y:a-histPred[i]}})).filter(d=>isFinite(d.x)&&isFinite(d.y));
new Chart(document.getElementById('residChart'),{{
  type:'scatter',
  data:{{datasets:[{{label:'Residuals',data:residData,backgroundColor:'rgba(59,130,246,0.4)',pointRadius:4}}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},
    scales:{{x:{{title:{{display:true,text:'Predicted',color:TXT}},ticks:{{}}  }},y:{{title:{{display:true,text:'Error',color:TXT}},ticks:{{}}}}  }}  }})
}});

// 7. TOP FEATURES BAR
new Chart(document.getElementById('topFeatChart'),{{
  type:'bar',
  data:{{labels:fiNames,datasets:[{{data:fiVals,
    backgroundColor:fiVals.map((_,i)=>`rgba(200,255,0,${{0.3+0.7*(i===0?1:fiVals[i]/fiVals[0])}})`),
    borderRadius:4}}]}},
  options:deepMerge(BASE,{{indexAxis:'y',plugins:{{legend:{{display:false}}}},
    scales:{{x:{{}},y:{{ticks:{{font:{{size:9}}}}}}  }}  }})
}});

// 8. PARETO CHART
const cumPct = fiVals.reduce((acc,v,i)=>{{
  const t=fiVals.reduce((a,b)=>a+b,0);
  acc.push(i===0?v/t*100:(acc[i-1]+v/t*100));
  return acc;
}},[]).map(v=>Math.round(v*10)/10);
new Chart(document.getElementById('paretoChart'),{{
  type:'bar',
  data:{{labels:fiNames,datasets:[
    {{label:'Importance',data:fiVals,backgroundColor:'rgba(59,130,246,0.6)',borderRadius:4,yAxisID:'y'}},
    {{label:'Cumulative %',data:cumPct,type:'line',borderColor:AC,pointRadius:4,
      pointBackgroundColor:AC,borderWidth:2,yAxisID:'y2'}}
  ]}},
  options:{{...BASE,plugins:{{legend:{{position:'top',labels:{{color:'#888',font:{{family:FONT,size:10}}}}}}}},
    scales:{{
      x:{{ticks:{{color:TXT,font:{{family:FONT,size:9}}}},grid:{{color:GRID}}}},
      y:{{ticks:{{color:TXT,font:{{family:FONT,size:9}}}},grid:{{color:GRID}},position:'left'}},
      y2:{{ticks:{{color:AC,font:{{family:FONT,size:9}},callback:v=>v+'%'}},grid:{{display:false}},position:'right',max:100}}
    }}  }}
}});

// 9. FORECAST BAR (dynamic period)
let fcChartRef = null;
function setFcPeriod(n,btn){{
  document.querySelectorAll('.fc-btn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('fcPeriodLabel').textContent=n+' Months';
  const labels=foreDates.slice(0,n), vals=foreVals.slice(0,n);
  if(fcChartRef)fcChartRef.destroy();
  fcChartRef=new Chart(document.getElementById('fcBarChart'),{{
    type:'bar',
    data:{{labels,datasets:[{{data:vals,
      backgroundColor:vals.map(v=>v>=avgSales?'rgba(200,255,0,0.7)':'rgba(239,68,68,0.6)'),
      borderRadius:4}}]}},
    options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},
      scales:{{x:{{ticks:{{maxRotation:45}}}},y:{{}}}}  }})
  }});
}}
setFcPeriod(3, document.querySelector('.fc-btn.active'));

// 10. MoM CHART
new Chart(document.getElementById('momChart'),{{
  type:'bar',
  data:{{labels:momDates,datasets:[{{
    label:'MoM Growth %',data:momVals,
    backgroundColor:momVals.map(v=>v>=0?'rgba(34,197,94,0.6)':'rgba(239,68,68,0.6)'),
    borderRadius:3
  }}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},
    scales:{{x:{{ticks:{{maxTicksLimit:12}}}},y:{{ticks:{{callback:v=>v+'%'}}}}  }}  }})
}});

// 11. QUARTERLY
new Chart(document.getElementById('quarterChart'),{{
  type:'bar',
  data:{{labels:qLabels,datasets:[{{data:qVals,
    backgroundColor:'rgba(249,115,22,0.6)',borderRadius:4}}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},scales:{{x:{{}},y:{{}}}}  }})
}});

// 12. MODEL BAR
const mNames=Object.keys(modelComp).map(n=>n.toUpperCase());
const mAccs=Object.values(modelComp).map(m=>Math.max(0,100-(m.mape||100)));
new Chart(document.getElementById('modelBarChart'),{{
  type:'bar',
  data:{{labels:mNames,datasets:[{{data:mAccs,
    backgroundColor:Object.keys(modelComp).map(n=>n===bestModel?AC:'#2a2a2a'),
    borderColor:Object.keys(modelComp).map(n=>n===bestModel?AC:'#333'),
    borderWidth:1,borderRadius:4}}]}},
  options:deepMerge(BASE,{{plugins:{{legend:{{display:false}}}},
    scales:{{x:{{}},y:{{min:0,max:105,ticks:{{callback:v=>v+'%'}}}}  }}  }})
}});

// 13. FEATURE IMPORTANCE BAR (models tab)
new Chart(document.getElementById('fiChart'),{{
  type:'bar',
  data:{{labels:fiNames,datasets:[{{data:fiVals,
    backgroundColor:fiVals.map((_,i)=>`rgba(200,255,0,${{0.25+0.75*(i===0?1:fiVals[i]/(fiVals[0]||1))}})`),
    borderRadius:4}}]}},
  options:deepMerge(BASE,{{indexAxis:'y',plugins:{{legend:{{display:false}}}},
    scales:{{x:{{}},y:{{ticks:{{font:{{size:9}}}}}}  }}  }})
}});

// ── SEASONAL HEATMAP ──────────────────────────────────────────────────────────
(function(){{
  const grid = document.getElementById('heatmapGrid');
  const lbls = document.getElementById('heatmapLabels');
  const max = Math.max(...seasonalAvg)||1;
  const min = Math.min(...seasonalAvg);
  seasonalAvg.forEach((v,i)=>{{
    const pct = (v-min)/(max-min+1e-10);
    const r = Math.round(pct*200), g = Math.round(pct*255);
    const cell = document.createElement('div');
    cell.className='heat-cell';
    cell.style.background=`rgba(${{r}},${{g}},0,0.7)`;
    cell.style.color=pct>0.5?'#000':'#fff';
    cell.textContent=v.toLocaleString(undefined,{{maximumFractionDigits:0}});
    grid.appendChild(cell);
    const lbl=document.createElement('div');
    lbl.className='heat-label';lbl.textContent=MONTH_NAMES[i];
    lbls.appendChild(lbl);
  }});
}})();

// ── MODEL TABLE ───────────────────────────────────────────────────────────────
(function(){{
  const tbody=document.getElementById('modelTbody');
  Object.entries(modelComp).forEach(([name,m])=>{{
    const isBest=name===bestModel;
    const acc=Math.max(0,100-(m.mape||100));
    tbody.innerHTML+=`<tr class="${{isBest?'best-row':''}}">
      <td style="font-weight:700">${{name.toUpperCase()}}</td>
      <td style="color:${{acc>90?GR:acc>75?OR:RD}};font-weight:700">${{acc.toFixed(1)}}%</td>
      <td>${{(m.mae||0).toFixed(4)}}</td>
      <td>${{(m.rmse||0).toFixed(4)}}</td>
      <td>${{(m.r2||0).toFixed(4)}}</td>
      <td>${{(m.mape||0).toFixed(2)}}%</td>
      <td style="width:120px"><div class="prog ${{isBest?'':'blue'}}"><div class="prog-fill" style="width:${{Math.round(acc)}}%"></div></div></td>
      <td><span class="badge ${{isBest?'badge-a':'badge-b'}}">${{isBest?'BEST':'TESTED'}}</span></td>
    </tr>`;
  }});
}})();

// ── FORECAST TABLE ────────────────────────────────────────────────────────────
(function(){{
  const tb=document.getElementById('fcTbody');
  foreDates.forEach((d,i)=>{{
    const v=foreVals[i]||0;
    const diff=v-avgSales, pct=avgSales>0?(diff/avgSales*100).toFixed(1):'0.0';
    const col=diff>=0?GR:RD;
    const bar=Math.round(Math.min(100,v/((Math.max(...foreVals)||1))*100));
    tb.innerHTML+=`<tr>
      <td style="color:var(--muted)">${{i+1}}</td>
      <td style="font-weight:700">${{d}}</td>
      <td style="font-weight:700">${{v.toLocaleString(undefined,{{maximumFractionDigits:0}})}}</td>
      <td style="color:${{col}}">${{diff>=0?'+':''}}${{diff.toFixed(0)}}</td>
      <td style="color:${{col}}">${{diff>=0?'+':''}}${{pct}}%</td>
      <td style="font-size:14px;color:${{col}}">${{diff>=0?'▲':'▼'}}</td>
      <td style="width:90px"><div class="prog ${{diff>=0?'green':'red'}}"><div class="prog-fill" style="width:${{bar}}%"></div></div></td>
    </tr>`;
  }});
}})();

// ── INSIGHTS ──────────────────────────────────────────────────────────────────
(function(){{
  const grid=document.getElementById('insightsGrid');
  const colors=['','green','blue','orange','red',''];
  insights.forEach((txt,i)=>{{
    const div=document.createElement('div');
    div.className='insight '+(colors[i%colors.length]||'');
    div.innerHTML=txt;
    grid.appendChild(div);
  }});
}})();

// ── DRILLTHROUGH ─────────────────────────────────────────────────────────────
(function(){{
  const grid=document.getElementById('drillGrid');
  let drillChartRef=null;
  fiNames.slice(0,8).forEach((name,i)=>{{
    const card=document.createElement('div');
    card.className='drill-card';
    card.innerHTML=`<div class="drill-name">${{name}}</div>
      <div class="drill-val">${{fiVals[i].toFixed(4)}}</div>
      <div class="drill-sub">Importance rank #${{i+1}}</div>`;
    card.onclick=()=>{{
      document.querySelectorAll('.drill-card').forEach(c=>c.classList.remove('selected'));
      card.classList.add('selected');
      document.getElementById('drillDetail').style.display='grid';
      document.getElementById('drillName').textContent=name;
      const pct=fiVals[0]>0?(fiVals[i]/fiVals[0]*100).toFixed(1):'0';
      const relToAvg=(fiVals[i]-fiVals.reduce((a,b)=>a+b,0)/fiVals.length);
      document.getElementById('drillStats').innerHTML=`
        <div class="mini-kpi"><span class="mini-kpi-label">Importance Score</span><span class="mini-kpi-val" style="color:var(--accent)">${{fiVals[i].toFixed(6)}}</span></div>
        <div class="mini-kpi"><span class="mini-kpi-label">Rank</span><span class="mini-kpi-val">#${{i+1}} of ${{fiNames.length}}</span></div>
        <div class="mini-kpi"><span class="mini-kpi-label">% of Top Feature</span><span class="mini-kpi-val">${{pct}}%</span></div>
        <div class="mini-kpi"><span class="mini-kpi-label">vs Average</span><span class="mini-kpi-val" style="color:${{relToAvg>=0?GR:RD}}">${{relToAvg>=0?'+':''}}${{relToAvg.toFixed(4)}}</span></div>
        <div class="mini-kpi"><span class="mini-kpi-label">Category</span><span class="mini-kpi-val">${{i<3?'Critical':i<6?'Important':'Supporting'}}</span></div>`;
      if(drillChartRef)drillChartRef.destroy();
      const ctx=document.getElementById('drillChart');
      drillChartRef=new Chart(ctx,{{
        type:'bar',
        data:{{labels:fiNames,datasets:[{{data:fiVals,
          backgroundColor:fiNames.map((_,j)=>j===i?AC:'#2a2a2a'),
          borderColor:fiNames.map((_,j)=>j===i?AC:'#333'),
          borderWidth:1,borderRadius:4}}]}},
        options:deepMerge(BASE,{{indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{}},y:{{ticks:{{font:{{size:8}}}}}}  }}  }})
      }});
    }};
    grid.appendChild(card);
  }});
}})();

// ── RECOMMENDATIONS ───────────────────────────────────────────────────────────
(function(){{
  const acc={accuracy:.1f}, gr={gr:.2f};
  const recs=[
    gr>0?'Continue current growth strategy — historical trend is positive.':'Investigate root cause of declining trend and model corrective action.',
    acc>90?'Model accuracy is excellent — forecast is highly reliable for planning.':'Collect more historical data to improve model accuracy.',
    'Monitor MoM growth chart for early signs of trend reversal.',
    'Use the 3-month forecast for short-term procurement/staffing decisions.',
    'Re-train the model monthly as new sales data becomes available.',
    'Top feature \\"'+fiNames[0]+'\\" should be tracked as a leading indicator.',
  ];
  const ul=document.getElementById('recList');
  recs.forEach(r=>{{ const li=document.createElement('li'); li.innerHTML=r; ul.appendChild(li); }});
}})();

// ── TABS ──────────────────────────────────────────────────────────────────────
function showTab(name){{
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.nbtn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+name).classList.add('active');
  document.getElementById('panel-'+name).classList.add('active');
}}

// ── EXPORT ────────────────────────────────────────────────────────────────────
function saveHTML(){{
  const b=new Blob([document.documentElement.outerHTML],{{type:'text/html;charset=utf-8'}});
  const a=document.createElement('a'); a.href=URL.createObjectURL(b);
  a.download='reviq_sales_report.html'; a.click();
}}
function exportCSV(){{
  let csv='Period,Date,Forecast,vs_Avg,vs_Avg_Pct\\n';
  foreDates.forEach((d,i)=>{{
    const v=foreVals[i]||0, diff=v-avgSales;
    csv+=`${{i+1}},${{d}},${{v.toFixed(2)}},${{diff.toFixed(2)}},${{avgSales>0?(diff/avgSales*100).toFixed(1):'0.0'}}\\n`;
  }});
  const b=new Blob([csv],{{type:'text/csv'}});
  const a=document.createElement('a'); a.href=URL.createObjectURL(b);
  a.download='reviq_forecast.csv'; a.click();
}}
</script>
</body>
</html>"""

        html_path = os.path.join(self.output_dir, f"{prefix}_dashboard.html")
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return html_path

    def _create_powerbi_readme(self, files: Dict[str, str], prefix: str) -> str:
        pbit_file = os.path.basename(files.get('pbit', f'{prefix}_dashboard.pbit'))
        hist_csv = os.path.basename(files.get('historical_csv', f'{prefix}_historical.csv'))
        fore_csv = os.path.basename(files.get('forecast_csv', f'{prefix}_forecast.csv'))
        html_file = os.path.basename(files.get('html_dashboard', f'{prefix}_dashboard.html'))

        readme = f"""# AI Sales Forecasting - Power BI Dashboard

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Files Generated
- `{html_file}` - **Interactive HTML Dashboard** (opens instantly in any browser)
- `{pbit_file}` - **Power BI Template** (import into Power BI Desktop)
- `{hist_csv}` - Historical sales data with predictions
- `{fore_csv}` - 12-month sales forecast
- `{os.path.basename(files.get('model_comparison_csv',''))}` - Model comparison metrics
- `{os.path.basename(files.get('feature_importance_csv',''))}` - Feature importance

## How to Open in Power BI Desktop

### Option A: Use HTML Dashboard (Recommended - Works Immediately)
1. Double-click `{html_file}` -> Opens in browser instantly
2. Click "Save Dashboard" button to save a copy

### Option B: Power BI Desktop (.pbit Template)
1. Install Power BI Desktop: https://powerbi.microsoft.com/desktop
2. Double-click `{pbit_file}` OR open Power BI Desktop -> File -> Import -> Power BI template
3. Point to the CSV files when prompted

### Option C: Import CSVs Manually
1. Open Power BI Desktop
2. Home -> Get Data -> Text/CSV
3. Import `{hist_csv}` and `{fore_csv}`
4. Create visuals using the data

### Option D: Power BI Service (Cloud)
1. Go to app.powerbi.com
2. Create new report -> Upload CSV files
3. Build visuals from the data

## Saving / Downloading from Power BI
- In Power BI Desktop: File -> Save As -> .pbix file
- In Power BI Service: File -> Download -> .pbix

## Data Schema
- `actual`: Real historical sales values
- `predicted`: ML model predictions
- `date`: Time period
- `forecast_date` + `forecast_value`: Future predictions
"""
        readme_path = os.path.join(self.output_dir, "README.md")
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write(readme)
        return readme_path

    def open_dashboard(self, html_path: str) -> bool:
        """Launch Power BI Desktop or fall back to browser"""
        # Try Power BI Desktop first
        pbi_paths = [
            r"C:\Program Files\Microsoft Power BI Desktop\bin\PBIDesktop.exe",
            r"C:\Program Files (x86)\Microsoft Power BI Desktop\bin\PBIDesktop.exe",
            "/Applications/Microsoft Power BI Desktop.app/Contents/MacOS/Microsoft Power BI Desktop",
        ]
        for path in pbi_paths:
            if os.path.exists(path):
                try:
                    subprocess.Popen([path])
                    logger.info("Power BI Desktop launched")
                    return True
                except Exception:
                    pass

        # Fall back to browser
        try:
            abs_path = os.path.abspath(html_path)
            webbrowser.open(f"file://{abs_path}")
            logger.info(f"Dashboard opened in browser: {abs_path}")
            return True
        except Exception as e:
            logger.error(f"Could not open dashboard: {e}")
            return False